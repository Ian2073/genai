#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Integrated Report Generator
Combines all reporting functions: terminal analysis, Excel reports, and visualization charts
"""

import json
import logging
import os
import sys
import statistics
import math
from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
import numpy as np

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from shared.stats_utils import pearson_correlation as calculate_correlation
from shared.stats_utils import spearman_correlation as calculate_spearman
from shared.score_utils import extract_raw_score as _extract_raw_score
from shared.score_utils import normalize_score_fields
from shared.story_data import load_story_records
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure font support for charts
matplotlib.use("Agg")
import platform
system = platform.system()
if system == "Windows":
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'DejaVu Sans', 'Arial']
elif system == "Darwin":  # macOS
    plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'Heiti TC', 'DejaVu Sans', 'Arial']
else:  # Linux
    plt.rcParams['font.sans-serif'] = ['WenQuanYi Micro Hei', 'DejaVu Sans', 'Arial']
plt.rcParams['axes.unicode_minus'] = False


logger = logging.getLogger(__name__)

# ===== Excel Tool: Auto Column Width =====
def _string_display_width(text: str) -> int:
    """Estimate the display width of a string in Excel.

    - ASCII characters count as 1
    - CJK full-width characters count as approximately 2
    - Returns at least 1
    """
    if text is None:
        return 1
    s = str(text)
    width = 0
    for ch in s:
        # Rough detection of CJK range
        if '\u4e00' <= ch <= '\u9fff' or '\u3040' <= ch <= '\u30ff' or '\uac00' <= ch <= '\ud7a3':
            width += 2
        else:
            width += 1
    return max(width, 1)


def _autofit_columns(worksheet, min_width: int = 8, max_width: int = 80, padding: int = 2) -> None:
    """Auto-adjust column widths based on content, for openpyxl worksheets.

    min_width/max_width control bounds; padding adds left/right margins.
    """
    from openpyxl.utils import get_column_letter

    # Collect maximum width for each column
    max_widths = {}
    for row in worksheet.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            current = _string_display_width(value)
            if idx not in max_widths:
                max_widths[idx] = current
            else:
                if current > max_widths[idx]:
                    max_widths[idx] = current

    # Also consider headers (avoid truncation)
    if worksheet.max_row >= 1:
        for idx, cell in enumerate(worksheet[1], start=1):
            header_w = _string_display_width(cell.value)
            max_widths[idx] = max(max_widths.get(idx, 1), header_w)

    # Actually set column widths
    for col_idx, w in max_widths.items():
        adjusted = min(max(w + padding, min_width), max_width)
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = adjusted

# Delayed import of evaluator for reusing alignment logic when needed
try:
    from evaluator import MultiAspectEvaluator  # Only for recalculating alignment scores, won't load large models
except Exception:
    MultiAspectEvaluator = None  # If import fails, disable alignment score recalculation

def load_user_ratings(roots=None):
    """Load user rating data for all stories"""
    roots = roots or ['output']
    user_ratings = {}
    records = load_story_records(roots, require_metadata=True)

    for record in records:
        data = record.get('metadata') or {}
        story_name = record.get('story_name')
        if not story_name:
            continue
        user_ratings[story_name] = {
            'user_rating': data.get('user_rating', None),
            'ratings_count': data.get('ratings_count', None),
            'platform': data.get('platform', 'Unknown'),
            'title': data.get('title', story_name)
        }
    
    return user_ratings

def load_assessment_reports(reapply_latest: bool = False, roots=None):
    """Load all assessment reports and integrate user rating data"""
    roots = roots or ['output']
    user_ratings = load_user_ratings(roots)
    reports = []
    evaluator = None
    if reapply_latest and MultiAspectEvaluator is not None:
        try:
            evaluator = MultiAspectEvaluator(
                enable_parallel_processing=False,
                preload_all_models=False,
                batch_size_optimization=False
            )
        except Exception:
            evaluator = None
    records = load_story_records(roots, require_report=True)

    for record in records:
        data = normalize_score_fields(dict(record.get('report') or {}))
        story_name = record.get('story_name')
        if not story_name:
            continue

        data['story_name'] = story_name

        # 整合用戶評價數據
        if story_name in user_ratings:
            data['user_rating'] = user_ratings[story_name]['user_rating']
            data['ratings_count'] = user_ratings[story_name]['ratings_count']
            data['platform'] = user_ratings[story_name]['platform']
            data['original_title'] = user_ratings[story_name]['title']
        else:
            data['user_rating'] = None
            data['ratings_count'] = None
            data['platform'] = 'Unknown'
            data['original_title'] = story_name

        # 選擇性：使用最新對齊模型重新計算 Cal 分（不改寫檔案，只在報告內使用）
        if reapply_latest and evaluator is not None:
            try:
                story_dir = record.get('story_dir') or ''
                # 取 base 分數（raw）
                raw = _extract_raw_score(data)
                if not isinstance(raw, (int, float)):
                    raw = data.get('original_overall_score') or data.get('overall_score')

                metadata = record.get('metadata') or {}
                evaluator.current_story_metadata = metadata
                # 調用對齊
                new_cal, _ = evaluator._apply_human_alignment(data.get('dimension_scores') or {}, float(raw))
                # 在記憶體中覆蓋用於分析與圖表的分數
                data['overall_score'] = float(round(new_cal, 1))
                data['overall_score_calibrated'] = float(round(new_cal, 1))
                data['score'] = dict(data.get('score') or {}, aligned=float(round(new_cal, 1)))
            except Exception:
                # 不中斷報告生成
                pass

        reports.append(data)
    
    return reports

def get_correlation_strength(correlation):
    """Determine correlation strength based on correlation coefficient"""
    abs_corr = abs(correlation)
    if abs_corr >= 0.7:
        return "strong"
    elif abs_corr >= 0.5:
        return "moderate"
    elif abs_corr >= 0.3:
        return "weak"
    else:
        return "very weak"


def _compute_governance_summary(reports):
    """彙整 governance 監控指標。"""
    risk_counts = {}
    review_counts = {}
    confidence_values = []
    high_risk_cases = []

    for report in reports:
        governance = report.get('governance') or {}
        if not isinstance(governance, dict) or not governance:
            continue

        risk = str(governance.get('risk_level') or 'unknown')
        risk_counts[risk] = risk_counts.get(risk, 0) + 1

        review = str(governance.get('review_recommendation') or 'unknown')
        review_counts[review] = review_counts.get(review, 0) + 1

        conf = governance.get('confidence')
        if isinstance(conf, (int, float)):
            confidence_values.append(float(conf))

        if risk in ('high', 'critical'):
            flags = governance.get('risk_flags') or []
            first_flag = ''
            if isinstance(flags, list) and flags:
                item = flags[0]
                if isinstance(item, dict):
                    first_flag = str(item.get('code') or '')
            high_risk_cases.append({
                'story': report.get('story_name'),
                'overall': report.get('overall_score'),
                'risk': risk,
                'review': review,
                'flag': first_flag,
            })

    high_risk_cases.sort(key=lambda x: float(x.get('overall') or 0.0), reverse=True)
    return {
        'risk_counts': risk_counts,
        'review_counts': review_counts,
        'confidence_mean': round(statistics.mean(confidence_values), 4) if confidence_values else None,
        'confidence_median': round(statistics.median(confidence_values), 4) if confidence_values else None,
        'high_risk_cases': high_risk_cases[:15],
    }

def print_terminal_analysis(reports):
    """Display analysis results in terminal"""
    # Sort by overall score
    reports.sort(key=lambda x: x['overall_score'], reverse=True)

    dimension_labels = {
        'coherence': ('Coher', 'Coherence'),
        'readability': ('Read', 'Readability'),
        'completeness': ('Compl', 'Completeness'),
        'entity_consistency': ('Entity', 'Entity Consistency'),
        'factuality': ('Fact', 'Factuality'),
        'emotional_impact': ('Emot', 'Emotional Impact')
    }

    logger.info('=== All Story Assessment Results (with User Rating Comparison) ===')
    for i, report in enumerate(reports, 1):
        scores = report['dimension_scores']
        story_name = report['story_name']
        overall = report['overall_score']
        raw = (report.get('score') or {}).get('base', report.get('overall_score_raw'))
        cal = (report.get('score') or {}).get('aligned', report.get('overall_score_calibrated'))

        dimension_parts = []
        for dim_key, (short_label, _) in dimension_labels.items():
            value = scores.get(dim_key)
            if isinstance(value, (int, float)):
                dimension_parts.append(f"{short_label}: {value:5.1f}")

        # User rating information
        user_rating = report.get('user_rating', 'N/A')
        ratings_count = report.get('ratings_count', 'N/A')

        if user_rating != 'N/A' and user_rating is not None:
            user_rating_str = f"{user_rating:.2f}"
            ratings_str = f"({ratings_count} ratings)" if ratings_count != 'N/A' else ""
            # Additional Raw/Cal vs User difference
            user_100 = user_rating * 20
            raw_seg = f" | Raw: {raw:5.1f}" if isinstance(raw, (int, float)) else ""
            cal_seg = f" | Cal: {cal:5.1f}" if isinstance(cal, (int, float)) else ""
            diff_raw = f" | Raw-User:{(raw - user_100):+5.1f}" if isinstance(raw, (int, float)) else ""
            diff_cal = f" | Cal-User:{(cal - user_100):+5.1f}" if isinstance(cal, (int, float)) else ""
            rating_info = f" | User: {user_rating_str}{ratings_str}{raw_seg}{cal_seg}{diff_raw}{diff_cal}"
        else:
            rating_info = " | User: No data"

        dimensions_summary = " | ".join(dimension_parts)
        detail_segment = f" | {dimensions_summary}" if dimensions_summary else ""
        logger.info("%2d. %-30s | AI Score: %5.1f%s%s", i, story_name, overall, detail_segment, rating_info)

    # User rating vs AI assessment correlation analysis - show specific comparison cases first
    valid_reports = [r for r in reports if r.get('user_rating') is not None and r.get('user_rating') != 'N/A']
    
    if len(valid_reports) >= 2:
        logger.info("")
        logger.info('=== Detailed Comparison Cases (Raw/Cal) ===')
        for report in valid_reports:
            story_name = report['story_name']
            raw = _extract_raw_score(report)
            cal = report.get('overall_score')
            user_rating = report['user_rating']
            ratings_count = report.get('ratings_count', 'N/A')
            user_score_100 = user_rating * 20
            diff_raw = (raw - user_score_100) if isinstance(raw, (int, float)) else None
            diff_cal = (cal - user_score_100) if isinstance(cal, (int, float)) else None
            logger.info('%-30s | Raw: %5.1f | Cal: %5.1f | User: %.2f(%.1f) | Raw-User:%+5.1f | Cal-User:%+5.1f | Ratings: %s',
                        story_name,
                        raw if isinstance(raw, (int, float)) else float('nan'),
                        cal if isinstance(cal, (int, float)) else float('nan'),
                        user_rating,
                        user_score_100,
                        diff_raw if diff_raw is not None else float('nan'),
                        diff_cal if diff_cal is not None else float('nan'),
                        ratings_count)
    
    logger.info("")
    logger.info('Total Stories Assessed: %s', len(reports))

    # Calculate statistics
    overall_scores = [r['overall_score'] for r in reports if isinstance(r.get('overall_score'), (int, float))]

    logger.info("")
    logger.info('=== Statistical Summary ===')
    if overall_scores:
        mean_overall = statistics.mean(overall_scores)
        std_overall = statistics.stdev(overall_scores) if len(overall_scores) > 1 else 0.0
        logger.info('Average Score: %.1f', mean_overall)
        logger.info('Highest Score: %.1f', max(overall_scores))
        logger.info('Lowest Score: %.1f', min(overall_scores))
        logger.info('Standard Deviation: %.1f', std_overall)
    else:
        logger.info('Missing valid overall score data')

    logger.info("")
    logger.info('=== Dimension Statistics ===')
    for dim_key, (_, full_label) in dimension_labels.items():
        dim_scores = [r['dimension_scores'].get(dim_key) for r in reports
                      if isinstance(r['dimension_scores'].get(dim_key), (int, float))]
        if not dim_scores:
            continue
        mean_dim = statistics.mean(dim_scores)
        std_dim = statistics.stdev(dim_scores) if len(dim_scores) > 1 else 0.0
        logger.info('%-20s: Avg %5.1f | Max %5.1f | Min %5.1f | Std %5.1f',
                    full_label,
                    mean_dim,
                    max(dim_scores),
                    min(dim_scores),
                    std_dim)
    
    # User rating vs AI assessment correlation analysis
    if len(valid_reports) >= 2:
        user_ratings = [r['user_rating'] for r in valid_reports]
        ai_scores_cal = [r['overall_score'] for r in valid_reports]
        ai_scores_raw = [_extract_raw_score(r) for r in valid_reports]
        # Calculate correlation coefficients (Raw and Cal)
        correlation_cal = calculate_correlation(ai_scores_cal, user_ratings)
        correlation_raw = calculate_correlation(ai_scores_raw, user_ratings)
        spearman_cal = calculate_spearman(ai_scores_cal, user_ratings)
        spearman_raw = calculate_spearman(ai_scores_raw, user_ratings)

        logger.info("")
        logger.info('=== User Rating vs AI Assessment Correlation Analysis ===')
        logger.info('Valid Stories with Data: %s', len(valid_reports))
        logger.info('Cal vs User Pearson r: %.3f | Spearman ρ: %.3f', correlation_cal, spearman_cal)
        logger.info('Raw vs User Pearson r: %.3f | Spearman ρ: %.3f', correlation_raw, spearman_raw)

        correlation_strength = get_correlation_strength(correlation_cal)
        direction = "positive" if correlation_cal > 0 else "negative"
        logger.info('Correlation Strength (based on Cal): %s %s correlation', correlation_strength, direction)

        logger.info("")
        # Correlation of each dimension with user ratings (unweighted)
        logger.info('=== Dimension vs User Rating Correlation (Unweighted) ===')
        for dim_key, (_, full_label) in dimension_labels.items():
            dim_vals = [r['dimension_scores'].get(dim_key) for r in valid_reports]
            paired = [(d, u) for d, u in zip(dim_vals, user_ratings) if isinstance(d, (int, float))]
            if len(paired) >= 2:
                dim_corr = calculate_correlation([p[0] for p in paired], [p[1] for p in paired])
                logger.info('%-20s: Correlation %.3f', full_label, dim_corr)
    else:
        logger.info("")
        logger.info('Insufficient user rating data for correlation analysis')

    governance_summary = _compute_governance_summary(reports)
    logger.info("")
    logger.info('=== Governance Routing Summary ===')
    logger.info('Risk Distribution: %s', governance_summary.get('risk_counts') or {})
    logger.info('Review Distribution: %s', governance_summary.get('review_counts') or {})
    logger.info('Confidence Mean/Median: %s / %s',
                governance_summary.get('confidence_mean'),
                governance_summary.get('confidence_median'))
    high_risk_cases = governance_summary.get('high_risk_cases') or []
    if high_risk_cases:
        logger.info('Top High/Critical Cases:')
        for case in high_risk_cases[:5]:
            logger.info('  - %s | score=%s | risk=%s | review=%s | flag=%s',
                        case.get('story'), case.get('overall'), case.get('risk'),
                        case.get('review'), case.get('flag'))

def create_excel_report(reports, report_dir="reports"):
    """Create professional Excel format report"""
    # Sort by overall score
    reports.sort(key=lambda x: x['overall_score'], reverse=True)
    
    dimension_config = [
        ('coherence', 'Coherence'),
        ('readability', 'Readability'),
        ('completeness', 'Completeness'),
        ('entity_consistency', 'Entity Consistency'),
        ('factuality', 'Factuality'),
        ('emotional_impact', 'Emotional Impact')
    ]

    def format_score(value):
        return round(float(value), 1) if isinstance(value, (int, float)) else 'N/A'

    # Create main data table
    main_data = []
    for i, report in enumerate(reports, 1):
        scores = report['dimension_scores']

        # User rating data
        user_rating = report.get('user_rating')
        ratings_count = report.get('ratings_count')
        platform = report.get('platform', 'Unknown')

        # Calculate user rating converted to 100-point scale
        raw = _extract_raw_score(report)
        cal = report.get('overall_score')
        user_score_100 = user_rating * 20 if user_rating is not None else None
        diff_raw = (raw - user_score_100) if (user_score_100 is not None and isinstance(raw, (int, float))) else None
        diff_cal = (cal - user_score_100) if user_score_100 is not None else None

        row = {
            'Rank': i,
            'Story Name': report['story_name'],
            'Raw Score': round(raw, 1) if isinstance(raw, (int, float)) else 'N/A',
            'Calibrated Score': round(cal, 1) if isinstance(cal, (int, float)) else 'N/A',
            'User Rating': round(user_rating, 2) if user_rating is not None else 'N/A',
            'User Rating (100-scale)': round(user_score_100, 1) if user_score_100 is not None else 'N/A',
            'Raw-User Diff': round(diff_raw, 1) if diff_raw is not None else 'N/A',
            'Cal-User Diff': round(diff_cal, 1) if diff_cal is not None else 'N/A',
            'Ratings Count': ratings_count if ratings_count is not None else 'N/A',
            'Platform': platform
        }

        for dim_key, display_name in dimension_config:
            row[display_name] = format_score(scores.get(dim_key))

        row['Processing Time (s)'] = round(report['processing_summary']['total_processing_time'], 1)
        main_data.append(row)

    # Create statistical data
    overall_scores = [r['overall_score'] for r in reports if isinstance(r.get('overall_score'), (int, float))]

    # Score distribution statistics
    score_ranges = [
        (75, 80, "Excellent"),
        (70, 74, "Good"), 
        (60, 69, "Average"),
        (50, 59, "Needs Improvement")
    ]

    distribution_data = []
    if overall_scores:
        total_reports = len(overall_scores)
        for min_score, max_score, level in score_ranges:
            count = sum(1 for score in overall_scores if min_score <= score <= max_score)
            percentage = round(count / total_reports * 100, 1)
            example = next((r['story_name'] for r in reports if min_score <= r['overall_score'] <= max_score), "") if count > 0 else ""
            distribution_data.append({
                'Score Range': f'{min_score}-{max_score}',
                'Level': level,
                'Story Count': count,
                'Percentage': f'{percentage}%',
                'Example': example
            })

    # Dimension statistics
    dimensions_data = []
    for dim_key, display_name in dimension_config:
        dim_scores = [r['dimension_scores'].get(dim_key) for r in reports
                      if isinstance(r['dimension_scores'].get(dim_key), (int, float))]
        if not dim_scores:
            continue
        avg_score = round(statistics.mean(dim_scores), 1)
        max_score = round(max(dim_scores), 1)
        min_score = round(min(dim_scores), 1)
        std_score = round(statistics.pstdev(dim_scores), 1) if len(dim_scores) > 1 else 0.0
        dimensions_data.append({
            'Dimension': display_name,
            'Average': avg_score,
            'Maximum': max_score,
            'Minimum': min_score,
            'Std Dev': std_score
        })
    
    # Create Excel file
    wb = Workbook()
    wb.remove(wb.active)

    governance_summary = _compute_governance_summary(reports)
    
    # Create main data sheet
    ws_main = wb.create_sheet("All Story Assessment Results")
    df_main = pd.DataFrame(main_data)
    
    for r in dataframe_to_rows(df_main, index=False, header=True):
        ws_main.append(r)
    
    # Set main table format
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Set header row format (three simple steps: font, background, center)
    for cell in ws_main[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Set all data rows to center alignment
    for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row):
        for cell in row:
            cell.alignment = center_alignment
    
    # Auto-adjust column widths (avoid truncating long titles)
    _autofit_columns(ws_main, min_width=8, max_width=80, padding=2)
    
    # Create score distribution table (interval statistics: Excellent/Good/Average/Needs Improvement)
    ws_dist = wb.create_sheet("Score Distribution")
    df_dist = pd.DataFrame(distribution_data)
    
    for r in dataframe_to_rows(df_dist, index=False, header=True):
        ws_dist.append(r)
    
    # Set distribution table format (bold and colored headers, centered content)
    for cell in ws_dist[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Set distribution table data rows to center alignment
    for row in ws_dist.iter_rows(min_row=2, max_row=ws_dist.max_row):
        for cell in row:
            cell.alignment = center_alignment
    
    # Auto-adjust column widths
    _autofit_columns(ws_dist, min_width=10, max_width=60, padding=2)
    
    # Create dimension statistics table (for each dimension: Average/Max/Min/Std Dev)
    ws_dims = wb.create_sheet("Dimension Performance")
    df_dims = pd.DataFrame(dimensions_data)
    
    for r in dataframe_to_rows(df_dims, index=False, header=True):
        ws_dims.append(r)
    
    # Set dimension table format (bold and colored headers, centered content)
    for cell in ws_dims[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Set dimension table data rows to center alignment
    for row in ws_dims.iter_rows(min_row=2, max_row=ws_dims.max_row):
        for cell in row:
            cell.alignment = center_alignment
    
    # Auto-adjust column widths
    _autofit_columns(ws_dims, min_width=10, max_width=60, padding=2)
    
    # Create user rating correlation analysis table (AI vs User score relationship)
    ws_corr = wb.create_sheet("User Rating Correlation")
    
    # Calculate correlation data
    valid_reports = [r for r in reports if r.get('user_rating') is not None and r.get('user_rating') != 'N/A']
    
    if len(valid_reports) >= 2:
        user_ratings = [r['user_rating'] for r in valid_reports]
        ai_scores_cal = [r['overall_score'] for r in valid_reports]
        ai_scores_raw = [_extract_raw_score(r) for r in valid_reports]
        # Calculate correlation coefficients (Raw and Cal)
        correlation_cal = calculate_correlation(ai_scores_cal, user_ratings)
        correlation_raw = calculate_correlation(ai_scores_raw, user_ratings)
        spearman_cal = calculate_spearman(ai_scores_cal, user_ratings)
        spearman_raw = calculate_spearman(ai_scores_raw, user_ratings)
        
        # Create correlation analysis data
        correlation_data = [
            {'Metric': 'Valid Stories Count', 'Value': len(valid_reports), 'Description': 'Stories with user rating data'},
            {'Metric': 'Cal vs User Pearson r', 'Value': round(correlation_cal, 3), 'Description': 'Calibrated vs user Pearson correlation'},
            {'Metric': 'Cal vs User Spearman ρ', 'Value': round(spearman_cal, 3), 'Description': 'Calibrated vs user rank correlation'},
            {'Metric': 'Raw vs User Pearson r', 'Value': round(correlation_raw, 3), 'Description': 'Uncalibrated vs user Pearson correlation'},
            {'Metric': 'Raw vs User Spearman ρ', 'Value': round(spearman_raw, 3), 'Description': 'Uncalibrated vs user rank correlation'},
            {'Metric': 'Correlation Strength (Cal)', 'Value': get_correlation_strength(correlation_cal), 'Description': 'Based on Cal correlation coefficient'}
        ]
        
        # Add detailed comparison data
        comparison_data = []
        for report in valid_reports:
            user_score_100 = report['user_rating'] * 20
            raw = _extract_raw_score(report)
            cal = report.get('overall_score')
            diff_raw = raw - user_score_100 if isinstance(raw, (int, float)) else None
            diff_cal = cal - user_score_100 if isinstance(cal, (int, float)) else None
            comparison_data.append({
                'Story Name': report['story_name'],
                'Raw Score': round(raw, 1) if isinstance(raw, (int, float)) else 'N/A',
                'Calibrated Score': round(cal, 1) if isinstance(cal, (int, float)) else 'N/A',
                'User Rating': round(report['user_rating'], 2),
                'User Rating (100-scale)': round(user_score_100, 1),
                'Raw-User Diff': round(diff_raw, 1) if diff_raw is not None else 'N/A',
                'Cal-User Diff': round(diff_cal, 1) if diff_cal is not None else 'N/A',
                'Ratings Count': report.get('ratings_count', 'N/A'),
                'Platform': report.get('platform', 'Unknown')
            })
        
        # Write correlation analysis data
        df_corr = pd.DataFrame(correlation_data)
        for r in dataframe_to_rows(df_corr, index=False, header=True):
            ws_corr.append(r)
        
        # Set format
        for cell in ws_corr[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Set correlation analysis table data rows to center alignment
        for row in ws_corr.iter_rows(min_row=2, max_row=len(correlation_data) + 1):
            for cell in row:
                cell.alignment = center_alignment
        
        # Auto-adjust column widths (correlation summary)
        _autofit_columns(ws_corr, min_width=10, max_width=80, padding=2)
        
        # Add blank row
        ws_corr.append([])
        ws_corr.append(['Detailed Comparison Data'])
        
        # Write comparison data (list AI vs user differences for each story)
        df_comp = pd.DataFrame(comparison_data)
        for r in dataframe_to_rows(df_comp, index=False, header=True):
            ws_corr.append(r)
        
        # Set comparison data format
        header_row = len(correlation_data) + 3  # Correlation data rows + blank row + title row
        for cell in ws_corr[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Set comparison data rows to center alignment
        for row in ws_corr.iter_rows(min_row=header_row + 1, max_row=ws_corr.max_row):
            for cell in row:
                cell.alignment = center_alignment
        
        # Auto-adjust again, covering the entire worksheet (including comparison section)
        _autofit_columns(ws_corr, min_width=10, max_width=80, padding=2)
    else:
        ws_corr.append(['Insufficient user rating data for correlation analysis'])

    ws_gov = wb.create_sheet("Governance Overview")
    governance_rows = [
        {
            'Metric': 'Confidence Mean',
            'Value': governance_summary.get('confidence_mean'),
            'Description': 'Average governance confidence across reports',
        },
        {
            'Metric': 'Confidence Median',
            'Value': governance_summary.get('confidence_median'),
            'Description': 'Median governance confidence across reports',
        },
    ]

    for risk, count in sorted((governance_summary.get('risk_counts') or {}).items()):
        governance_rows.append({
            'Metric': f'Risk Count: {risk}',
            'Value': count,
            'Description': 'Number of reports at this risk level',
        })

    for review, count in sorted((governance_summary.get('review_counts') or {}).items()):
        governance_rows.append({
            'Metric': f'Review Route: {review}',
            'Value': count,
            'Description': 'Number of reports routed to this review action',
        })

    df_gov = pd.DataFrame(governance_rows)
    for r in dataframe_to_rows(df_gov, index=False, header=True):
        ws_gov.append(r)

    for cell in ws_gov[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for row in ws_gov.iter_rows(min_row=2, max_row=ws_gov.max_row):
        for cell in row:
            cell.alignment = center_alignment

    ws_gov.append([])
    ws_gov.append(['High/Critical Cases'])
    header_row = ws_gov.max_row + 1

    case_rows = governance_summary.get('high_risk_cases') or []
    if case_rows:
        df_cases = pd.DataFrame(case_rows)
        for r in dataframe_to_rows(df_cases, index=False, header=True):
            ws_gov.append(r)
        for cell in ws_gov[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        for row in ws_gov.iter_rows(min_row=header_row + 1, max_row=ws_gov.max_row):
            for cell in row:
                cell.alignment = center_alignment

    _autofit_columns(ws_gov, min_width=10, max_width=80, padding=2)
    
    # Save file to specified folder
    os.makedirs(report_dir, exist_ok=True)
    report_path = os.path.join(report_dir, "AI_Assessment_Report.xlsx")
    
    # If file is in use, add timestamp
    if os.path.exists(report_path):
        try:
            wb.save(report_path)
        except PermissionError:
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            report_path = os.path.join(report_dir, f"AI_Assessment_Report_{timestamp}.xlsx")
            wb.save(report_path)
    else:
        wb.save(report_path)
    
    # Excel report generated (path will be shown in final summary)

def create_visualizations(reports, report_dir="reports"):
    """Create visualization charts"""
    # Collect stories with user rating data
    valid_data = []
    for report in reports:
        story_name = report['story_name']
        if report.get('user_rating') is not None:
            ai_score = report['overall_score']
            user_rating = report['user_rating']
            ratings_count = report.get('ratings_count', 0)
            platform = report.get('platform', 'Unknown')
            
            valid_data.append({
                'story_name': story_name,
                'ai_score': ai_score,
                'user_rating': user_rating,
                'user_score_100': user_rating * 20,  # 轉換為百分制
                'ratings_count': ratings_count,
                'platform': platform
            })
    
    if len(valid_data) < 2:
        logger.warning("Insufficient user rating data to generate visualization charts")
        return
    
    # Ensure report folder exists
    os.makedirs(report_dir, exist_ok=True)
    
    # 1. Scatter plot
    ai_scores = [d['ai_score'] for d in valid_data]
    user_scores_100 = [d['user_score_100'] for d in valid_data]
    story_names = [d['story_name'] for d in valid_data]
    ratings_counts = [d['ratings_count'] for d in valid_data]
    
    correlation = calculate_correlation(ai_scores, user_scores_100)
    
    fig, ax = plt.subplots(figsize=(16, 10))
    # Uniform size, use color to represent rating count
    uniform_size = 150  # Uniform point size
    
    # If rating count range is large, use logarithmic normalization to improve color distribution
    if ratings_counts:
        min_count = min(ratings_counts)
        max_count = max(ratings_counts)
        
        # If range is large, use logarithmic normalization
        if max_count / min_count > 10:
            # Use logarithmically normalized values to map colors
            log_counts = [math.log10(count) for count in ratings_counts]
            scatter = ax.scatter(ai_scores, user_scores_100, s=uniform_size, alpha=0.7, 
                               c=log_counts, cmap='viridis', edgecolors='black', linewidths=0.8,
                               norm=plt.Normalize(vmin=min(log_counts), vmax=max(log_counts)))
        else:
            scatter = ax.scatter(ai_scores, user_scores_100, s=uniform_size, alpha=0.7, 
                               c=ratings_counts, cmap='viridis', edgecolors='black', linewidths=0.8)
    else:
        scatter = ax.scatter(ai_scores, user_scores_100, s=uniform_size, alpha=0.7, 
                           c='steelblue', edgecolors='black', linewidths=0.8)
    
    # Add trend line
    z = np.polyfit(ai_scores, user_scores_100, 1)
    p = np.poly1d(z)
    ax.plot(ai_scores, p(ai_scores), "r--", alpha=0.8, linewidth=2)
    
    ax.set_xlabel('AI Assessment Score', fontsize=14)
    ax.set_ylabel('User Rating Score (100-point scale)', fontsize=14)
    ax.set_title(f'AI Assessment Score vs User Rating Score\nCorrelation: {correlation:.3f}', fontsize=16, fontweight='bold', pad=20)
    ax.grid(True, alpha=0.3)
    
    # Add color bar showing rating count, use better tick spacing
    cbar = plt.colorbar(scatter, ax=ax)
    cbar.set_label('Ratings Count', fontsize=12, labelpad=10)
    
    # 調整顏色條刻度，使用更合理的分佈
    if ratings_counts:
        min_count = min(ratings_counts)
        max_count = max(ratings_counts)
        
        # 如果評價數範圍很大，使用對數刻度
        if max_count / min_count > 10:
            from matplotlib.ticker import FuncFormatter
            # 設置對數刻度位置
            log_min = math.log10(min_count)
            log_max = math.log10(max_count)
            num_ticks = 6
            log_ticks = np.logspace(log_min, log_max, num_ticks)
            # 將刻度值轉換回對數空間
            log_ticks_log = [math.log10(tick) for tick in log_ticks]
            cbar.set_ticks(log_ticks_log)
            # 格式化刻度標籤，顯示原始數值（使用千分位格式）
            def format_func(x, p):
                # 將對數值轉回原始值
                original_val = 10 ** x
                if original_val >= 1000:
                    return f'{original_val/1000:.0f}K'
                return f'{original_val:.0f}'
            cbar.ax.yaxis.set_major_formatter(FuncFormatter(format_func))
        else:
            # 範圍較小，使用線性刻度
            from matplotlib.ticker import FuncFormatter
            num_ticks = 6
            linear_ticks = np.linspace(min_count, max_count, num_ticks)
            cbar.set_ticks(linear_ticks)
            # 格式化刻度標籤
            def format_func(x, p):
                if x >= 1000:
                    return f'{x/1000:.1f}K'
                return f'{x:.0f}'
            cbar.ax.yaxis.set_major_formatter(FuncFormatter(format_func))
    
    correlation_strength = get_correlation_strength(correlation)
    direction = "正" if correlation > 0 else "負"
    ax.text(0.05, 0.95, f'相關性強度: {correlation_strength}{direction}相關', 
            transform=ax.transAxes, fontsize=12, 
            bbox=dict(boxstyle="round,pad=0.3", facecolor="lightblue", alpha=0.7),
            verticalalignment='top')
    
    plt.tight_layout(pad=3.0)
    scatter_path = os.path.join(report_dir, 'AI評估vs用戶評價_散點圖.png')
    plt.savefig(scatter_path, dpi=300, bbox_inches='tight', pad_inches=0.5)
    plt.close()
    
    # 2. 對比柱狀圖 - 顯示所有數據，數據量大時自動分頁
    # 計算差異並排序
    for i, data in enumerate(valid_data):
        data['diff'] = data['ai_score'] - data['user_score_100']
    
    # 按AI分數排序
    valid_data.sort(key=lambda x: x['ai_score'], reverse=True)
    
    num_stories = len(valid_data)
    stories_per_page = 50  # 每頁顯示50個故事
    
    # 如果數據量大於每頁顯示數量，分成多頁
    if num_stories > stories_per_page:
        num_pages = (num_stories + stories_per_page - 1) // stories_per_page  # 向上取整
        for page in range(num_pages):
            start_idx = page * stories_per_page
            end_idx = min((page + 1) * stories_per_page, num_stories)
            page_data = valid_data[start_idx:end_idx]
            
            story_names = [d['story_name'] for d in page_data]
            ai_scores = [d['ai_score'] for d in page_data]
            user_scores = [d['user_score_100'] for d in page_data]
            differences = [d['diff'] for d in page_data]
            
            num_display = len(story_names)
            fig_width = max(20, num_display * 0.5)
            fig_height = 10
            
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))
            x = np.arange(num_display)
            width = 0.35
            
            bars1 = ax.bar(x - width/2, ai_scores, width, label='AI評估分數', color='skyblue', alpha=0.8)
            bars2 = ax.bar(x + width/2, user_scores, width, label='用戶評價分數', color='lightcoral', alpha=0.8)
            
            ax.set_xlabel('故事名稱', fontsize=14)
            ax.set_ylabel('分數', fontsize=14)
            ax.set_title(f'AI評估分數 vs 用戶評價分數對比（第{page+1}頁，共{num_pages}頁）', 
                        fontsize=16, fontweight='bold', pad=20)
            ax.set_xticks(x)
            ax.set_xticklabels(story_names, rotation=90, ha='center', va='top', fontsize=9)
            
            ax.legend(fontsize=12, loc='upper right')
            ax.grid(True, alpha=0.3, axis='y')
            
            # 只對差異較大的案例顯示差異標籤
            for i, diff in enumerate(differences):
                if abs(diff) > 5:
                    color = 'green' if diff > 0 else 'red' if diff < 0 else 'gray'
                    max_score = max(ai_scores[i], user_scores[i])
                    y_pos = max_score + 1
                    ax.text(i, y_pos, f'{diff:+.1f}', 
                            ha='center', va='bottom', fontsize=8, color=color, fontweight='bold',
                            bbox=dict(boxstyle='round,pad=0.15', facecolor='white', alpha=0.8, edgecolor=color, linewidth=0.8))
            
            plt.tight_layout(pad=3.0)
            plt.subplots_adjust(bottom=0.2)
            filename = os.path.join(report_dir, f'AI評估vs用戶評價_對比圖_第{page+1}頁.png')
            plt.savefig(filename, dpi=300, bbox_inches='tight', pad_inches=0.5)
            plt.close()
    else:
        # 數據量小，單頁顯示
        story_names = [d['story_name'] for d in valid_data]
        ai_scores = [d['ai_score'] for d in valid_data]
        user_scores = [d['user_score_100'] for d in valid_data]
        differences = [d['diff'] for d in valid_data]
        
        fig_width = max(20, num_stories * 0.5)
        fig_height = 10
        
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))
        x = np.arange(num_stories)
        width = 0.35
        
        bars1 = ax.bar(x - width/2, ai_scores, width, label='AI評估分數', color='skyblue', alpha=0.8)
        bars2 = ax.bar(x + width/2, user_scores, width, label='用戶評價分數', color='lightcoral', alpha=0.8)
        
        ax.set_xlabel('故事名稱', fontsize=14)
        ax.set_ylabel('分數', fontsize=14)
        ax.set_title('AI評估分數 vs 用戶評價分數對比', fontsize=16, fontweight='bold', pad=20)
        ax.set_xticks(x)
        ax.set_xticklabels(story_names, rotation=90, ha='center', va='top', fontsize=9)
        
        ax.legend(fontsize=12, loc='upper right')
        ax.grid(True, alpha=0.3, axis='y')
        
        # 只對差異較大的案例顯示差異標籤
        for i, diff in enumerate(differences):
            if abs(diff) > 5:
                color = 'green' if diff > 0 else 'red' if diff < 0 else 'gray'
                max_score = max(ai_scores[i], user_scores[i])
                y_pos = max_score + 1
                ax.text(i, y_pos, f'{diff:+.1f}', 
                        ha='center', va='bottom', fontsize=8, color=color, fontweight='bold',
                        bbox=dict(boxstyle='round,pad=0.15', facecolor='white', alpha=0.8, edgecolor=color, linewidth=0.8))
        
        plt.tight_layout(pad=3.0)
        plt.subplots_adjust(bottom=0.2)
        compare_path = os.path.join(report_dir, 'AI評估vs用戶評價_對比圖.png')
        plt.savefig(compare_path, dpi=300, bbox_inches='tight', pad_inches=0.5)
        plt.close()
    
    # 3. 差異分析圖 - 顯示所有數據，數據量大時自動分頁
    # 使用已計算的 diff 字段，按差異排序
    display_data = sorted(valid_data, key=lambda x: x.get('diff', x['ai_score'] - x['user_score_100']), reverse=True)
    
    num_stories = len(display_data)
    stories_per_page = 50  # 每頁顯示50個故事
    
    # 如果數據量大於每頁顯示數量，分成多頁
    if num_stories > stories_per_page:
        num_pages = (num_stories + stories_per_page - 1) // stories_per_page
        for page in range(num_pages):
            start_idx = page * stories_per_page
            end_idx = min((page + 1) * stories_per_page, num_stories)
            page_data = display_data[start_idx:end_idx]
            
            story_names = [d['story_name'] for d in page_data]
            differences = [d.get('diff', d['ai_score'] - d['user_score_100']) for d in page_data]
            
            num_display = len(story_names)
            fig_height = max(12, num_display * 0.3)
            fig_width = 16
            
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))
            colors = ['green' if d > 0 else 'red' if d < 0 else 'gray' for d in differences]
            
            bars = ax.barh(story_names, differences, color=colors, alpha=0.7, height=0.6)
            
            ax.set_xlabel('評分差異 (AI分數 - 用戶分數)', fontsize=14, labelpad=10)
            ax.set_ylabel('故事名稱', fontsize=14, labelpad=10)
            ax.set_title(f'AI評估與用戶評價的評分差異分析（第{page+1}頁，共{num_pages}頁）', 
                        fontsize=16, fontweight='bold', pad=20)
            ax.grid(True, alpha=0.3, axis='x')
            
            y_label_size = max(9, min(11, 12 - num_display * 0.02))
            ax.tick_params(axis='y', labelsize=y_label_size, pad=6)
            ax.tick_params(axis='x', labelsize=12, pad=8)
            
            ax.set_ylim(-0.5, len(story_names) - 0.5)
            ax.axvline(x=0, color='black', linestyle='-', linewidth=1.5, alpha=0.6)
            
            # 只對差異較大的案例顯示數值標籤
            for i, (bar, diff) in enumerate(zip(bars, differences)):
                width = bar.get_width()
                if abs(width) > 5:
                    offset = 0.5 if width >= 0 else -0.5
                    ax.text(width + offset, bar.get_y() + bar.get_height()/2, 
                            f'{diff:+.1f}', ha='left' if width >= 0 else 'right', va='center', 
                            fontsize=8, fontweight='bold',
                            bbox=dict(boxstyle='round,pad=0.15', facecolor='white', alpha=0.8, edgecolor='none'))
            
            from matplotlib.patches import Patch
            legend_elements = [Patch(facecolor='green', alpha=0.7, label='AI評分高於用戶'),
                              Patch(facecolor='red', alpha=0.7, label='AI評分低於用戶')]
            ax.legend(handles=legend_elements, loc='lower right', fontsize=12, framealpha=0.9)
            
            max_name_length = max(len(name) for name in story_names) if story_names else 20
            left_margin = max(0.25, 0.18 + max(0, (max_name_length - 20) * 0.008))
            plt.subplots_adjust(left=left_margin, right=0.95, top=0.95, bottom=0.05)
            plt.tight_layout(pad=3.0)
            filename = os.path.join(report_dir, f'評分差異分析圖_第{page+1}頁.png')
            plt.savefig(filename, dpi=300, bbox_inches='tight', pad_inches=0.5)
            plt.close()
    else:
        # 數據量小，單頁顯示
        story_names = [d['story_name'] for d in display_data]
        differences = [d.get('diff', d['ai_score'] - d['user_score_100']) for d in display_data]
        
        fig_height = max(12, num_stories * 0.3)
        fig_width = 16
        
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))
        colors = ['green' if d > 0 else 'red' if d < 0 else 'gray' for d in differences]
        
        bars = ax.barh(story_names, differences, color=colors, alpha=0.7, height=0.6)
        
        ax.set_xlabel('評分差異 (AI分數 - 用戶分數)', fontsize=14, labelpad=10)
        ax.set_ylabel('故事名稱', fontsize=14, labelpad=10)
        ax.set_title('AI評估與用戶評價的評分差異分析', fontsize=16, fontweight='bold', pad=20)
        ax.grid(True, alpha=0.3, axis='x')
        
        y_label_size = max(7, min(10, 12 - num_stories * 0.01))
        ax.tick_params(axis='y', labelsize=y_label_size, pad=6)
        ax.tick_params(axis='x', labelsize=12, pad=8)
        
        ax.set_ylim(-0.5, len(story_names) - 0.5)
        ax.axvline(x=0, color='black', linestyle='-', linewidth=1.5, alpha=0.6)
        
        for i, (bar, diff) in enumerate(zip(bars, differences)):
            width = bar.get_width()
            if abs(width) > 5:
                offset = 0.5 if width >= 0 else -0.5
                ax.text(width + offset, bar.get_y() + bar.get_height()/2, 
                        f'{diff:+.1f}', ha='left' if width >= 0 else 'right', va='center', 
                        fontsize=8, fontweight='bold',
                        bbox=dict(boxstyle='round,pad=0.15', facecolor='white', alpha=0.8, edgecolor='none'))
        
        from matplotlib.patches import Patch
        legend_elements = [Patch(facecolor='green', alpha=0.7, label='AI評分高於用戶'),
                          Patch(facecolor='red', alpha=0.7, label='AI評分低於用戶')]
        ax.legend(handles=legend_elements, loc='lower right', fontsize=12, framealpha=0.9)
        
        max_name_length = max(len(name) for name in story_names) if story_names else 20
        left_margin = max(0.25, 0.18 + max(0, (max_name_length - 20) * 0.008))
        plt.subplots_adjust(left=left_margin, right=0.95, top=0.95, bottom=0.05)
        plt.tight_layout(pad=3.0)
        diff_path = os.path.join(report_dir, '評分差異分析圖.png')
        plt.savefig(diff_path, dpi=300, bbox_inches='tight', pad_inches=0.5)
        plt.close()
    
    # 4. 統計圖表：箱線圖和殘差分析
    create_statistical_charts(valid_data, report_dir)

def create_statistical_charts(valid_data, report_dir="reports"):
    """創建統計分析圖表：箱線圖、殘差圖、分佈對比"""
    if len(valid_data) < 2:
        return
    
    ai_scores = [d['ai_score'] for d in valid_data]
    user_scores = [d['user_score_100'] for d in valid_data]
    differences = [d.get('diff', d['ai_score'] - d['user_score_100']) for d in valid_data]
    
    # 4.1 箱線圖：比較AI分數和用戶分數的分佈
    fig, ax = plt.subplots(figsize=(10, 8))
    box_data = [ai_scores, user_scores]
    box_labels = ['AI評估分數', '用戶評價分數']
    bp = ax.boxplot(box_data, tick_labels=box_labels, patch_artist=True, 
                    widths=0.6, showmeans=True, meanline=True)
    
    # 設置顏色
    colors_box = ['skyblue', 'lightcoral']
    for patch, color in zip(bp['boxes'], colors_box):
        patch.set_facecolor(color)
        patch.set_alpha(0.7)
    
    ax.set_ylabel('分數', fontsize=14)
    ax.set_title('AI評估分數 vs 用戶評價分數分佈對比（箱線圖）', fontsize=16, fontweight='bold', pad=20)
    ax.grid(True, alpha=0.3, axis='y')
    
    # 添加統計信息
    stats_text = f"AI評估: 平均={np.mean(ai_scores):.1f}, 中位數={np.median(ai_scores):.1f}\n"
    stats_text += f"用戶評價: 平均={np.mean(user_scores):.1f}, 中位數={np.median(user_scores):.1f}"
    ax.text(0.02, 0.98, stats_text, transform=ax.transAxes, fontsize=10,
            verticalalignment='top', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
    
    plt.tight_layout(pad=3.0)
    boxplot_path = os.path.join(report_dir, '分數分佈箱線圖.png')
    plt.savefig(boxplot_path, dpi=300, bbox_inches='tight', pad_inches=0.5)
    plt.close()
    
    # 4.2 殘差圖（Residual Plot）：顯示預測誤差的分佈
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))
    
    # 左圖：殘差 vs AI分數
    ax1.scatter(ai_scores, differences, alpha=0.6, s=50, color='steelblue', edgecolors='black', linewidths=0.5)
    ax1.axhline(y=0, color='red', linestyle='--', linewidth=2, alpha=0.7, label='零誤差線')
    ax1.set_xlabel('AI評估分數', fontsize=12)
    ax1.set_ylabel('殘差 (AI分數 - 用戶分數)', fontsize=12)
    ax1.set_title('殘差 vs AI評估分數', fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3)
    ax1.legend(fontsize=10)
    
    # 添加趨勢線
    if len(ai_scores) > 1:
        z = np.polyfit(ai_scores, differences, 1)
        p = np.poly1d(z)
        ax1.plot(ai_scores, p(ai_scores), "r-", alpha=0.5, linewidth=2, label=f'趨勢線: y={z[0]:.2f}x+{z[1]:.2f}')
        ax1.legend(fontsize=10)
    
    # 右圖：殘差直方圖
    ax2.hist(differences, bins=20, color='steelblue', alpha=0.7, edgecolor='black', linewidth=0.5)
    ax2.axvline(x=0, color='red', linestyle='--', linewidth=2, alpha=0.7, label='零誤差線')
    ax2.axvline(x=np.mean(differences), color='green', linestyle='-', linewidth=2, alpha=0.7, label=f'平均誤差: {np.mean(differences):.2f}')
    ax2.set_xlabel('殘差 (AI分數 - 用戶分數)', fontsize=12)
    ax2.set_ylabel('頻率', fontsize=12)
    ax2.set_title('殘差分佈直方圖', fontsize=14, fontweight='bold')
    ax2.grid(True, alpha=0.3, axis='y')
    ax2.legend(fontsize=10)
    
    plt.tight_layout(pad=3.0)
    residual_path = os.path.join(report_dir, '殘差分析圖.png')
    plt.savefig(residual_path, dpi=300, bbox_inches='tight', pad_inches=0.5)
    plt.close()
    
    # 4.3 分佈密度對比圖（KDE）
    try:
        from scipy import stats
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # 計算密度
        ai_density = stats.gaussian_kde(ai_scores)
        user_density = stats.gaussian_kde(user_scores)
        
        # 生成x軸範圍
        x_min = min(min(ai_scores), min(user_scores))
        x_max = max(max(ai_scores), max(user_scores))
        x_range = np.linspace(x_min - 5, x_max + 5, 200)
        
        # 繪製密度曲線
        ax.plot(x_range, ai_density(x_range), label='AI評估分數', linewidth=2.5, color='skyblue')
        ax.fill_between(x_range, ai_density(x_range), alpha=0.3, color='skyblue')
        
        ax.plot(x_range, user_density(x_range), label='用戶評價分數', linewidth=2.5, color='lightcoral')
        ax.fill_between(x_range, user_density(x_range), alpha=0.3, color='lightcoral')
        
        ax.set_xlabel('分數', fontsize=14)
        ax.set_ylabel('密度', fontsize=14)
        ax.set_title('AI評估分數 vs 用戶評價分數分佈密度對比', fontsize=16, fontweight='bold', pad=20)
        ax.legend(fontsize=12, loc='upper right')
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout(pad=3.0)
        density_path = os.path.join(report_dir, '分數分佈密度對比圖.png')
        plt.savefig(density_path, dpi=300, bbox_inches='tight', pad_inches=0.5)
        plt.close()
    except ImportError:
        logger.warning("scipy未安裝，跳過密度對比圖生成")
    except Exception as e:
        logger.warning("生成密度對比圖時出錯: %s", e)

def main():
    """主函數：生成所有報告"""
    # 設置終端編碼
    import sys
    import os
    import datetime
    import argparse
    if sys.platform.startswith('win'):
        os.system('chcp 65001 > nul')  # 設置UTF-8編碼
    
    # 解析命令列參數
    parser = argparse.ArgumentParser(description="生成統合報告（可選：在報告內重算最新對齊分）")
    parser.add_argument('--reapply-latest', action='store_true', help='在報告中使用最新對齊模型重算 Cal 分（不改寫原檔）')
    parser.add_argument('--roots', nargs='+', default=['output'], help='要掃描的故事根目錄（預設: output）')
    parser.add_argument('--output-dir', default='reports/evaluation', help='報告輸出根目錄（預設: reports/evaluation）')
    args = parser.parse_args()

    logger.info("開始生成統合報告...")
    
    # 讀取數據
    reports = load_assessment_reports(reapply_latest=bool(args.reapply_latest), roots=args.roots)
    
    if not reports:
        logger.warning("未找到評估報告數據")
        return
    
    # 創建報告資料夾（使用時間和故事數量命名，時間在前方便排序）
    time_str = datetime.datetime.now().strftime("%m%d_%H%M")
    num_stories = len(reports)
    report_dir = os.path.join(args.output_dir, f"{time_str}_{num_stories}個故事")
    os.makedirs(report_dir, exist_ok=True)
    
    logger.info("報告將保存至：%s", report_dir)
    
    # 1. 終端分析
    logger.info("")
    logger.info("%s", "=" * 60)
    logger.info("終端分析報告")
    logger.info("%s", "=" * 60)
    print_terminal_analysis(reports)
    
    # 2. Excel報告
    logger.info("")
    logger.info("%s", "=" * 60)
    logger.info("生成Excel報告")
    logger.info("%s", "=" * 60)
    create_excel_report(reports, report_dir)
    
    # 3. 視覺化圖表
    create_visualizations(reports, report_dir)
    
    logger.info("")
    logger.info("所有報告已生成完成：%s", report_dir)

if __name__ == '__main__':
    logging.basicConfig(
        level=getattr(logging, os.environ.get("REPORT_LOG_LEVEL", "INFO").upper(), logging.INFO),
        format="%(message)s"
    )
    main()
